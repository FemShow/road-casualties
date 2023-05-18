import openpyxl

# Load workbook
workbook = openpyxl.load_workbook("/Users/femisokoya/Documents/GitHub/road_length/road-casualties/road-casualties-severity-lsoa-msoa-ward.xlsx")

# Select worksheet
worksheet = workbook["LSOA11"]

# Get list of merged cells
merged_cells_copy = worksheet.merged_cells.ranges.copy()

# Unmerge all cells
for merged_cell in merged_cells_copy:
    worksheet.unmerge_cells(str(merged_cell))

# Loop through rows
for row in worksheet.iter_rows(min_row=1, max_row=1):
    # Loop through cells
    for cell in row:
        # If the cell is merged, write its value to all cells in the range
        if cell.coordinate in worksheet.merged_cells:
            cell_range = worksheet.merged_cells[cell.coordinate]
            cell_value = cell.value
            for merged_cell in cell_range:
                for merged_cell_row in range(merged_cell.min_row, merged_cell.max_row + 1):
                    for merged_cell_column in range(merged_cell.min_col, merged_cell.max_col + 1):
                        worksheet.cell(row=merged_cell_row, column=merged_cell_column).value = cell_value

        # If the cell has a value, insert a new column to the left with the header "Year"
        if cell.value:
            column_index = cell.column
            worksheet.insert_cols(column_index)
            worksheet.cell(row=2, column=column_index).value = "Year"
            for row_index in range(3, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).value = cell.value

# Delete the first row
worksheet.delete_rows(1)

# Save workbook
workbook.save("/Users/femisokoya/Documents/GitHub/road_length/road-casualties/example1.xlsx")
